import csv
import os
from openpyxl import Workbook

stud = dict()

map = {"AA": 10,"AB": 9,"BB": 8,"BC": 7,"CC": 6,"CD": 5,"DD": 4,"DD*": 4,"F": 0,"F*": 0,"I": 0}

course = dict()

def calc(grad, crd):
    avg = 0.0
    crd_sum = 0
    for i in range(len(grad)):
        avg += grad[i]*crd[i]
    for c in crd:
        crd_sum += c
    return round(avg/crd_sum, 2)

def generate_marksheet():
    if os.path.exists("output") == False:
        os.makedirs("output")

    with open("names-roll.csv", "r") as f:
        word = csv.DictReader(f)
        for i in word:
            stud[i["Roll"]] = {"Name": i["Name"]}
    with open("subjects_master.csv", "r") as f:
        word = csv.DictReader(f)
        for i in word:
            course[i["subno"]] = {"subname": i["subname"], "ltp": i["ltp"], "crd": i["crd"]}

    with open("grades.csv", "r") as f:
        word = csv.DictReader(f)
        for i in word:
            try:
                stud[i["Roll"]][i["Sem"]]
            except KeyError:
                stud[i["Roll"]][i["Sem"]] = {}
            stud[i["Roll"]][i["Sem"]][i["SubCode"]] = {"Grade": i["Grade"].strip(), "Sub_Type": i["Sub_Type"]}
    for i in stud:
        wb = Workbook()
        sem_cred = [0,0,0,0,0,0,0,0]
        tot_cred = [0,0,0,0,0,0,0,0]
        spi = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]
        cpi = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]
        for j in range(1, 9):
            cred = []
            grad = []
            ws = wb.create_sheet("Sem" + str(j))
            ws.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
            l = 1
            try:
                stud[i][str(j)]
            except KeyError:
                continue
            for k in stud[i][str(j)]:
                # print(k)
                # print(stud[i][str(j)][k])
                ws.append([l,k,course[k]["subname"],course[k]["ltp"],course[k]["crd"],stud[i][str(j)][k]["Sub_Type"],stud[i][str(j)][k]["Grade"]])
                cred.append(int(course[k]["crd"]))
                grad.append(map[stud[i][str(j)][k]["Grade"]])
                l += 1
            for c in cred:
                sem_cred[j-1]+=c
            spi[j-1]=calc(grad, cred)
            if j>1:
                tot_cred[j-1]=tot_cred[j-2]+sem_cred[j-1]
                cpi[j-1]=calc(spi[:j],sem_cred[:j])
            else:
                tot_cred[j-1]=sem_cred[j-1]
                cpi[j-1]=spi[j-1]
        ws = wb.create_sheet("Overall")
        ws.append(["Roll No.", i])
        ws.append(["Name of Student", stud[i]["Name"]])
        ws.append(["Discipline", i[4:6]])
        ws.append(["Semester No.", 1,2,3,4,5,6,7,8])
        tmp = ["Semester wise Credit Taken"]
        for j in sem_cred:
            tmp.append(str(j))
        ws.append(tmp)
        tmp = ["SPI"]
        for j in spi:
            tmp.append(str(j))
        ws.append(tmp)
        tmp = ["Total Credits Taken"]
        for j in tot_cred:
            tmp.append(str(j))
        ws.append(tmp)
        tmp = ["CPI"]
        for j in cpi:
            tmp.append(str(j))
        ws.append(tmp)
        wb.save(os.path.join("output", i + ".xlsx"))


generate_marksheet()
