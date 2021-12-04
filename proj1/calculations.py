import pandas as pd
import openpyxl
import os
import zipfile
import glob


def isNaN(num):
    return num != num

def op():
    read_file = pd.read_excel("./marksheets/concise_marksheet.xlsx")
    read_file.to_csv("./output/concise_marksheet.csv", index=None, header=True)

    def zip_directory(folder_path, zip_path):
        with zipfile.ZipFile(zip_path, mode='w') as zipf:
            len_dir_path = len(folder_path)
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, file_path[len_dir_path:])

    zip_directory('./marksheets', './output/marksheets.zip')
    return 1


def reset():

    os.chdir("./input/")
    for file in glob.glob("*"):
        os.remove(file)
    os.chdir("../")

    os.chdir("./marksheets/")
    for file in glob.glob("*"):
        os.remove(file)
    os.chdir("../")

    os.chdir("./output/")
    for file in glob.glob("*"):
        os.remove(file)
    os.chdir("../")
    return 1


def generate(name, pos, neg):
    pos = int(pos)
    neg = int(neg)
    mst = pd.read_csv('./input/master_roll.csv')
    stu = pd.read_csv('./input/responses.csv')
    x = list(stu["Roll Number"])

    col_nm = list(stu.columns)
    for i in range(0, len(col_nm)):
        if col_nm[i] == 'Roll Number':
            temp = col_nm[0]
            col_nm[0] = col_nm[i]
            col_nm[i] = temp
    stu = stu.reindex(columns=col_nm)

    for i in stu["Roll Number"]:
        if i == "ANSWER":
            break
        return "no answer"
    cmlist = list(stu.columns)
    cmlist.append('statsAns')
    cmlist.append('Score_After_Negative')

    cm_path = "marksheets/concise_marksheet.xlsx"
    wb = openpyxl.Workbook()
    cm = wb.active
    cm.append(cmlist)
    wb.save(cm_path)

    for i in mst["roll"]:
        if i == "ANSWER":
            temp = x.index(i)
            anskey = stu.loc[temp]

            ank = list(anskey)
            tempank = '[' + str(len(ank)-7) + ',0,0]'
            ank.append(tempank)
            tempank = str((len(ank)-8)*pos) + '/' + str((len(ank)-8)*pos)
            ank.append(tempank)

            cm_path = "marksheets/concise_marksheet.xlsx"
            wb = openpyxl.load_workbook(cm_path)
            cm = wb.active
            cm.append(ank)

            wb.save(cm_path)
            continue

        else:
            temp = x.index(i)
            stukey = stu.loc[temp]

            f_path = 'marksheets/' + i + '.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            thin = openpyxl.styles.Side(border_style="thin", color="000000")

            img = openpyxl.drawing.image.Image('./static/assets/img/ms.jpeg')
            wa = wb.worksheets[0]
            img.anchor = 'A1'
            wa.add_image(img)

            ws.merge_cells('A5:E5')
            cell = ws.cell(row=5, column=1)
            cell.value = "Mark Sheet"
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.font = openpyxl.styles.Font(name='Century', size=18, bold=True, underline='single')

            cell = ws.cell(row=6, column=1)
            cell.value = 'Name:'
            cell.font = openpyxl.styles.Font(name='Century', size=12)

            cell = ws.cell(row=6, column=2)
            cell.value = stukey['Name']
            cell.font = openpyxl.styles.Font(name='Century', size=12, bold=True)

            ws['A7'] = 'Roll Number:'
            ws['A7'].font = openpyxl.styles.Font(name='Century', size=12)
            ws['B7'] = stukey['Roll Number']
            ws['B7'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)

            ws['D6'] = 'Exam:'
            ws['D6'].font = openpyxl.styles.Font(name='Century', size=12)
            ws['E6'] = 'quiz'
            ws['E6'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)

            l1 = list(stukey[7:])
            l2 = list(anskey[7:])

            ws['A15'] = 'Student Ans'
            ws['A15'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['A15'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['B15'] = 'Correct Ans'
            ws['B15'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['B15'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            cnt1 = 25
            cnt2 = 16
            a = "A"
            b = "B"
            max = len(l2)
            maxmks = max * pos
            right = int(0)
            wrong = int(0)
            na = int(0)
            mks = int(0)
            for j in range(0, len(l2)):
                if cnt1 == 0:
                    a = "D"
                    b = "E"
                    cnt2 = 16
                    ws['D15'] = 'Student Ans'
                    ws['D15'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
                    ws['D15'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

                    ws['E15'] = 'Correct Ans'
                    ws['E15'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
                    ws['E15'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                if l1[j] == l2[j]:
                    s = a + str(cnt2)
                    ws[s] = l1[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='0000FF00')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    s = b + str(cnt2)
                    ws[s] = l2[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='000000FF')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    mks = mks + pos
                    right += 1
                    cnt2 += 1
                    cnt1 -= 1
                elif isNaN(l1[j]):
                    s = a + str(cnt2)
                    ws[s] = l1[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='000000FF')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    s = b + str(cnt2)
                    ws[s] = l2[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='000000FF')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    na += 1
                    cnt2 += 1
                    cnt1 -= 1
                elif l1[j] != l2[j]:
                    s = a + str(cnt2)
                    ws[s] = l1[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='00FF0000')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    s = b + str(cnt2)
                    ws[s] = l2[j]
                    ws[s].font = openpyxl.styles.Font(name='Century', size=12, color='000000FF')
                    ws[s].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    mks = mks - neg
                    wrong += 1
                    cnt2 += 1
                    cnt1 -= 1

            ws['B9'] = 'Right'
            ws['B9'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['B9'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['C9'] = 'Wrong'
            ws['C9'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['C9'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['D9'] = 'Not Attempt'
            ws['D9'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['D9'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['E9'] = 'Max.'
            ws['E9'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['E9'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['A9'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['A10'] = 'No.'
            ws['A10'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['A10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['A11'] = 'Marking'
            ws['A11'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['A11'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['A12'] = 'Total'
            ws['A12'].font = openpyxl.styles.Font(name='Century', size=12, bold=True)
            ws['A12'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['B10'] = right
            ws['B10'].font = openpyxl.styles.Font(name='Century', size=12, color='0000FF00')
            ws['B10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['B11'] = pos
            ws['B11'].font = openpyxl.styles.Font(name='Century', size=12, color='0000FF00')
            ws['B11'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['B12'] = right * pos
            ws['B12'].font = openpyxl.styles.Font(name='Century', size=12, color='0000FF00')
            ws['B12'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['C10'] = wrong
            ws['C10'].font = openpyxl.styles.Font(name='Century', size=12, color='00FF0000')
            ws['C10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['C11'] = neg
            ws['C11'].font = openpyxl.styles.Font(name='Century', size=12, color='00FF0000')
            ws['C11'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['C12'] = wrong * neg
            ws['C12'].font = openpyxl.styles.Font(name='Century', size=12, color='00FF0000')
            ws['C12'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['D10'] = na
            ws['D10'].font = openpyxl.styles.Font(name='Century', size=12)
            ws['D10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['D11'] = 0
            ws['D11'].font = openpyxl.styles.Font(name='Century', size=12)
            ws['D11'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['D12'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            ws['E10'] = max
            ws['E10'].font = openpyxl.styles.Font(name='Century', size=12)
            ws['E10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['E10'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['E12'] = str(mks) + '/' + str(maxmks)
            ws['E12'].font = openpyxl.styles.Font(name='Century', size=12, color='000000FF')
            ws['E12'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

            wb.save(f_path)

            stk = list(stukey)
            tempstk = '[' + str(right) + ',' + str(wrong) + ',' + str(na) + ']'
            stk.append(tempstk)
            tempstk = str(mks) + '/' + str(maxmks)
            stk.append(tempstk)

            cm_path = "marksheets/concise_marksheet.xlsx"
            wb = openpyxl.load_workbook(cm_path)
            cm = wb.active
            cm.append(stk)

            wb.save(cm_path)
    return "Success"